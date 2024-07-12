import openpyxl
import os
import shutil
from pyrevit import forms
from Autodesk.Revit.DB import PDFExportOptions


uidoc = __revit__.ActiveUIDocument
doc = __revit__.ActiveUIDocument.Document

#get hould of excel file using ironpython


FPath = forms.pick_file(file_ext='xlsx', multi_file=False, unc_paths=False)


clr.AddReference("Microsoft.Office.Interop.Excel")

import Microsoft.Office.Interop.Excel as Excel
excel = Excel.ApplicationClass()
excel.Visible = False
workbook = excel.Workbooks.Open(FPath)
xl = workbook.Worksheets['SheetDATA']

PDFExportOptions

pVal = xl.Cells(i,2).Value2



sourceFile ='C:\\Users\\Wagner.Human\\Desktop\\33808.00 - Gantries\\SANRAL Gantry Cantilever.xlsx'

wb = openpyxl.load_workbook('C:\\Users\\Wagner.Human\\Desktop\\33808.00 - Gantries\\R300 Table.xlsx')
sheet = wb['Sheet1'] # Get a sheet from the workbook.
sheet['A1'] # Get a cell from the sheet

i = 10
sheet[f'C{i}'].value # Get the value from the cell.


for i in range(1, 9):
    desFile = f'C:\\Users\\Wagner.Human\\Desktop\\33808.00 - Gantries\\SANRAL Gantry Cantilever{sheet[f"A{i}"].value}_{sheet[f"B{i}"].value}.xlsx'
    shutil.copy2(sourceFile,desFile)
    SRwb = openpyxl.load_workbook(desFile)
    SRsheet = SRwb['Geometry'] # Get a sheet from the workbook.
    SRsheet['F10'].value = sheet[f'E{i}'].value/1000
    SRsheet['F11'].value = sheet[f'F{i}'].value/1000
    SRwb.save(desFile)    
    print(f"run{i}")    




# wb = openpyxl.load_workbook('C:\\Users\\Wagner.Human\\Desktop\\33808.00 - Gantries\\SANRAL GANTRY PORTAL.xlsx')
# sheet = wb['Geometry'] # Get a sheet from the workbook.
# sheet['A1'] # Get a cell from the sheet

# i = 10
# sheet[f'C{i}'].value # Get the value from the cell.

# print(sheet[f'C{i}'].value)

# for i in range(10, 26):
    
#     sheet[f'C{i}'].value=sheet[f'A{count+i}'].value
        
        

# wb.save('example_copy.xlsx')